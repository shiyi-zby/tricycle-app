# 第一部分：窗口、鼠标配置，必须放在所有Kivy导入最前面
from kivy.config import Config
# 禁止窗口缩放
Config.set('kivy', 'window', 'resizable', '0')
Config.set('graphics', 'width', '360')
Config.set('graphics', 'height', '640')
# 关闭鼠标模拟多点触控（去掉桌面端红点，安卓不影响）
Config.set('input', 'mouse', 'mouse,disable_multitouch')

import os
import sys

# 打包资源路径兼容函数
def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

# 关闭日志弹窗、指定渲染后端，防止安卓闪退
os.environ['KIVY_NO_CONSOLELOG'] = '1'
os.environ['KIVY_GL_BACKEND'] = 'angle_sdl2'

from kivymd.app import MDApp
from kivymd.uix.screen import MDScreen
from kivymd.uix.dialog import MDDialog
from kivymd.uix.button import MDFlatButton, MDRaisedButton, MDFillRoundFlatButton, MDIconButton
from kivymd.uix.textfield import MDTextField
from kivymd.uix.card import MDCard
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.datatables import MDDataTable
from kivymd.uix.label import MDLabel
from kivymd.uix.list import MDList, TwoLineListItem, OneLineListItem
from kivymd.uix.menu import MDDropdownMenu
from kivy.uix.scrollview import ScrollView
from kivy.lang import Builder
from kivy.metrics import dp
from kivy.core.window import Window
from kivy.core.text import LabelBase
from datetime import datetime
import pandas as pd
import json

# 窗口尺寸适配
Window.size = (1050, 960)

# 界面布局KV
KV = '''
MDScreen:
    md_bg_color: 0.95, 0.95, 0.97, 1

    MDBoxLayout:
        orientation: 'vertical'
        spacing: dp(2)

        # 标题栏
        MDBoxLayout:
            size_hint_y: None
            height: dp(50)
            md_bg_color: app.theme_cls.primary_color
            padding: [dp(15), dp(5)]

            MDLabel:
                text: '三轮车配置录入(选择+手动)'
                halign: 'center'
                theme_text_color: 'Custom'
                text_color: 1, 1, 1, 1
                font_style: 'H6'
                font_name: "Roboto"
                bold: True

        # 统计栏
        MDCard:
            size_hint_y: None
            height: dp(40)
            padding: [dp(15), dp(5)]
            md_bg_color: 0.9, 0.95, 1, 1
            radius: [0, 0, 0, 0]
            elevation: 1

            MDLabel:
                id: stats_bar
                text: '已录入: 0个型号 | 总计: 0辆 | 总金额: ¥0'
                halign: 'center'
                theme_text_color: 'Primary'
                font_style: 'Caption'
                font_name: "Roboto"
                bold: True

        # 主内容滚动区
        ScrollView:
            do_scroll_x: False
            do_scroll_y: True
            bar_width: dp(4)

            MDBoxLayout:
                orientation: 'vertical'
                spacing: dp(8)
                size_hint_y: None
                height: self.minimum_height
                padding: [dp(10), dp(8), dp(10), dp(10)]

                # 车型号输入
                MDCard:
                    size_hint_y: None
                    height: dp(55)
                    padding: [dp(10), dp(5)]
                    elevation: 2
                    radius: [dp(8)]
                    md_bg_color: 1, 1, 1, 1

                    MDTextField:
                        id: model_input
                        hint_text: '车型号（手动输入）'
                        mode: 'rectangle'
                        size_hint_y: None
                        height: dp(45)
                        font_size: dp(14)

                # 配置选项区域
                MDCard:
                    orientation: 'vertical'
                    size_hint_y: None
                    height: dp(620)
                    padding: [dp(12), dp(8)]
                    elevation: 2
                    radius: [dp(8)]
                    md_bg_color: 1, 1, 1, 1
                    spacing: dp(3)

                    MDLabel:
                        text: '【配置选项】点击选择或点✏️手动输入'
                        size_hint_y: None
                        height: dp(25)
                        font_style: 'Subtitle2'
                        font_name: "Roboto"
                        bold: True
                        theme_text_color: 'Primary'
                        font_size: dp(11)

                    # 电机品牌
                    MDBoxLayout:
                        orientation: 'horizontal'
                        spacing: dp(5)
                        size_hint_y: None
                        height: dp(40)

                        MDLabel:
                            text: '电机品牌'
                            size_hint_x: 0.22
                            halign: 'right'
                            font_style: 'Body2'
                            font_name: "Roboto"
                            font_size: dp(12)

                        MDFillRoundFlatButton:
                            id: motor_brand_btn
                            text: '点击选择 ▼'
                            size_hint_x: 0.58
                            md_bg_color: 0.9, 0.9, 0.9, 1
                            text_color: 0, 0, 0, 1
                            font_size: dp(11)
                            on_release: app.open_menu('motor_brand')

                        MDIconButton:
                            icon: 'pencil'
                            size_hint_x: 0.15
                            icon_size: dp(18)
                            on_release: app.manual_input('motor_brand', '电机品牌')

                    # 电机功率
                    MDBoxLayout:
                        orientation: 'horizontal'
                        spacing: dp(5)
                        size_hint_y: None
                        height: dp(40)

                        MDLabel:
                            text: '电机功率'
                            size_hint_x: 0.22
                            halign: 'right'
                            font_style: 'Body2'
                            font_name: "Roboto"
                            font_size: dp(12)

                        MDFillRoundFlatButton:
                            id: motor_power_btn
                            text: '点击选择 ▼'
                            size_hint_x: 0.58
                            md_bg_color: 0.9, 0.9, 0.9, 1
                            text_color: 0, 0, 0, 1
                            font_size: dp(11)
                            on_release: app.open_menu('motor_power')

                        MDIconButton:
                            icon: 'pencil'
                            size_hint_x: 0.15
                            icon_size: dp(18)
                            on_release: app.manual_input('motor_power', '电机功率')

                    # 轮胎品牌
                    MDBoxLayout:
                        orientation: 'horizontal'
                        spacing: dp(5)
                        size_hint_y: None
                        height: dp(40)

                        MDLabel:
                            text: '轮胎品牌'
                            size_hint_x: 0.22
                            halign: 'right'
                            font_style: 'Body2'
                            font_name: "Roboto"
                            font_size: dp(12)

                        MDFillRoundFlatButton:
                            id: tire_brand_btn
                            text: '点击选择 ▼'
                            size_hint_x: 0.58
                            md_bg_color: 0.9, 0.9, 0.9, 1
                            text_color: 0, 0, 0, 1
                            font_size: dp(11)
                            on_release: app.open_menu('tire_brand')

                        MDIconButton:
                            icon: 'pencil'
                            size_hint_x: 0.15
                            icon_size: dp(18)
                            on_release: app.manual_input('tire_brand', '轮胎品牌')

                    # 前轮规格
                    MDBoxLayout:
                        orientation: 'horizontal'
                        spacing: dp(5)
                        size_hint_y: None
                        height: dp(40)

                        MDLabel:
                            text: '前轮规格'
                            size_hint_x: 0.22
                            halign: 'right'
                            font_style: 'Body2'
                            font_name: "Roboto"
                            font_size: dp(12)

                        MDFillRoundFlatButton:
                            id: front_tire_spec_btn
                            text: '点击选择 ▼'
                            size_hint_x: 0.58
                            md_bg_color: 0.9, 0.9, 0.9, 1
                            text_color: 0, 0, 0, 1
                            font_size: dp(11)
                            on_release: app.open_menu('front_tire_spec')

                        MDIconButton:
                            icon: 'pencil'
                            size_hint_x: 0.15
                            icon_size: dp(18)
                            on_release: app.manual_input('front_tire_spec', '前轮规格')

                    # 后轮规格
                    MDBoxLayout:
                        orientation: 'horizontal'
                        spacing: dp(5)
                        size_hint_y: None
                        height: dp(40)

                        MDLabel:
                            text: '后轮规格'
                            size_hint_x: 0.22
                            halign: 'right'
                            font_style: 'Body2'
                            font_name: "Roboto"
                            font_size: dp(12)

                        MDFillRoundFlatButton:
                            id: rear_tire_spec_btn
                            text: '点击选择 ▼'
                            size_hint_x: 0.58
                            md_bg_color: 0.9, 0.9, 0.9, 1
                            text_color: 0, 0, 0, 1
                            font_size: dp(11)
                            on_release: app.open_menu('rear_tire_spec')

                        MDIconButton:
                            icon: 'pencil'
                            size_hint_x: 0.15
                            icon_size: dp(18)
                            on_release: app.manual_input('rear_tire_spec', '后轮规格')

                    # 减震品牌
                    MDBoxLayout:
                        orientation: 'horizontal'
                        spacing: dp(5)
                        size_hint_y: None
                        height: dp(40)

                        MDLabel:
                            text: '减震品牌'
                            size_hint_x: 0.22
                            halign: 'right'
                            font_style: 'Body2'
                            font_name: "Roboto"
                            font_size: dp(12)

                        MDFillRoundFlatButton:
                            id: shock_brand_btn
                            text: '点击选择 ▼'
                            size_hint_x: 0.58
                            md_bg_color: 0.9, 0.9, 0.9, 1
                            text_color: 0, 0, 0, 1
                            font_size: dp(11)
                            on_release: app.open_menu('shock_brand')

                        MDIconButton:
                            icon: 'pencil'
                            size_hint_x: 0.15
                            icon_size: dp(18)
                            on_release: app.manual_input('shock_brand', '减震品牌')

                    # 减震型号
                    MDBoxLayout:
                        orientation: 'horizontal'
                        spacing: dp(5)
                        size_hint_y: None
                        height: dp(40)

                        MDLabel:
                            text: '减震型号'
                            size_hint_x: 0.22
                            halign: 'right'
                            font_style: 'Body2'
                            font_name: "Roboto"
                            font_size: dp(12)

                        MDFillRoundFlatButton:
                            id: shock_model_btn
                            text: '点击选择 ▼'
                            size_hint_x: 0.58
                            md_bg_color: 0.9, 0.9, 0.9, 1
                            text_color: 0, 0, 0, 1
                            font_size: dp(11)
                            on_release: app.open_menu('shock_model')

                        MDIconButton:
                            icon: 'pencil'
                            size_hint_x: 0.15
                            icon_size: dp(18)
                            on_release: app.manual_input('shock_model', '减震型号')

                    # 新增：后桥品牌
                    MDBoxLayout:
                        orientation: 'horizontal'
                        spacing: dp(5)
                        size_hint_y: None
                        height: dp(40)

                        MDLabel:
                            text: '后桥品牌'
                            size_hint_x: 0.22
                            halign: 'right'
                            font_style: 'Body2'
                            font_name: "Roboto"
                            font_size: dp(12)

                        MDFillRoundFlatButton:
                            id: rear_axle_brand_btn
                            text: '点击选择 ▼'
                            size_hint_x: 0.58
                            md_bg_color: 0.9, 0.9, 0.9, 1
                            text_color: 0, 0, 0, 1
                            font_size: dp(11)
                            on_release: app.open_menu('rear_axle_brand')

                        MDIconButton:
                            icon: 'pencil'
                            size_hint_x: 0.15
                            icon_size: dp(18)
                            on_release: app.manual_input('rear_axle_brand', '后桥品牌')

                    # 新增：后桥类型
                    MDBoxLayout:
                        orientation: 'horizontal'
                        spacing: dp(5)
                        size_hint_y: None
                        height: dp(40)

                        MDLabel:
                            text: '后桥类型'
                            size_hint_x: 0.22
                            halign: 'right'
                            font_style: 'Body2'
                            font_name: "Roboto"
                            font_size: dp(12)

                        MDFillRoundFlatButton:
                            id: rear_axle_type_btn
                            text: '点击选择 ▼'
                            size_hint_x: 0.58
                            md_bg_color: 0.9, 0.9, 0.9, 1
                            text_color: 0, 0, 0, 1
                            font_size: dp(11)
                            on_release: app.open_menu('rear_axle_type')

                        MDIconButton:
                            icon: 'pencil'
                            size_hint_x: 0.15
                            icon_size: dp(18)
                            on_release: app.manual_input('rear_axle_type', '后桥类型')

                    # 自定义选项管理
                    MDBoxLayout:
                        orientation: 'horizontal'
                        spacing: dp(5)
                        size_hint_y: None
                        height: dp(35)
                        padding: [0, dp(5), 0, 0]

                        MDRaisedButton:
                            text: '管理自定义选项'
                            md_bg_color: 0.6, 0.6, 0.6, 1
                            size_hint_x: 1
                            height: dp(35)
                            font_size: dp(11)
                            on_release: app.manage_custom_options()

                # 价格与数量区域
                MDCard:
                    orientation: 'vertical'
                    size_hint_y: None
                    height: dp(130)
                    padding: [dp(12), dp(8)]
                    elevation: 2
                    radius: [dp(8)]
                    md_bg_color: 1, 1, 1, 1
                    spacing: dp(5)

                    MDLabel:
                        text: '【价格与数量】'
                        size_hint_y: None
                        height: dp(25)
                        font_style: 'Subtitle2'
                        font_name: "Roboto"
                        bold: True
                        theme_text_color: 'Primary'

                    MDBoxLayout:
                        orientation: 'horizontal'
                        spacing: dp(8)
                        size_hint_y: None
                        height: dp(45)

                        MDTextField:
                            id: price_input
                            hint_text: '单价（元）*'
                            mode: 'rectangle'
                            input_filter: 'float'
                            size_hint_x: 0.5
                            height: dp(45)
                            font_size: dp(13)

                        MDTextField:
                            id: quantity_input
                            hint_text: '数量（辆）*'
                            mode: 'rectangle'
                            input_filter: 'int'
                            size_hint_x: 0.5
                            height: dp(45)
                            font_size: dp(13)

                    MDTextField:
                        id: remarks_input
                        hint_text: '备注（可选）'
                        mode: 'rectangle'
                        size_hint_y: None
                        height: dp(40)
                        font_size: dp(13)

                # 功能按钮区
                MDBoxLayout:
                    orientation: 'horizontal'
                    spacing: dp(8)
                    size_hint_y: None
                    height: dp(45)

                    MDRaisedButton:
                        text: '添加到列表'
                        md_bg_color: 0.2, 0.6, 0.9, 1
                        on_release: app.add_to_list()
                        size_hint_x: 0.5
                        height: dp(45)
                        font_size: dp(12)

                    MDRaisedButton:
                        text: '快速录入下一个'
                        md_bg_color: 0.3, 0.7, 0.3, 1
                        on_release: app.quick_add()
                        size_hint_x: 0.5
                        height: dp(45)
                        font_size: dp(12)

                MDBoxLayout:
                    orientation: 'horizontal'
                    spacing: dp(8)
                    size_hint_y: None
                    height: dp(45)

                    MDRaisedButton:
                        text: '查看完整表格'
                        md_bg_color: 0.5, 0.3, 0.7, 1
                        on_release: app.view_full_table()
                        size_hint_x: 0.5
                        height: dp(45)
                        font_size: dp(12)

                    MDRaisedButton:
                        text: '导出Excel'
                        md_bg_color: 0.8, 0.5, 0.1, 1
                        on_release: app.export_excel()
                        size_hint_x: 0.5
                        height: dp(45)
                        font_size: dp(12)

                # 已录入型号列表
                MDCard:
                    orientation: 'vertical'
                    size_hint_y: None
                    height: dp(200)
                    elevation: 2
                    radius: [dp(8)]
                    md_bg_color: 1, 1, 1, 1
                    padding: [dp(8), dp(5)]

                    MDBoxLayout:
                        orientation: 'horizontal'
                        size_hint_y: None
                        height: dp(30)
                        spacing: dp(5)

                        MDLabel:
                            text: '已录入型号列表：'
                            font_style: 'Subtitle2'
                            font_name: "Roboto"
                            bold: True
                            size_hint_x: 0.7

                        MDLabel:
                            id: count_label
                            text: '0个'
                            halign: 'right'
                            theme_text_color: 'Primary'
                            font_name: "Roboto"
                            size_hint_x: 0.3

                    ScrollView:
                        do_scroll_x: False
                        do_scroll_y: True

                        MDList:
                            id: model_list_preview
                            spacing: dp(2)

                # 数据管理按钮
                MDBoxLayout:
                    orientation: 'horizontal'
                    spacing: dp(8)
                    size_hint_y: None
                    height: dp(45)
                    padding: [0, dp(5), 0, dp(10)]

                    MDRaisedButton:
                        text: '删除选中'
                        md_bg_color: 0.95, 0.5, 0.1, 1
                        on_release: app.delete_selected()
                        size_hint_x: 0.33
                        height: dp(45)
                        font_size: dp(11)

                    MDRaisedButton:
                        text: '删除最后'
                        md_bg_color: 0.9, 0.3, 0.1, 1
                        on_release: app.delete_last()
                        size_hint_x: 0.33
                        height: dp(45)
                        font_size: dp(11)

                    MDRaisedButton:
                        text: '清空全部'
                        md_bg_color: 0.8, 0.1, 0.1, 1
                        on_release: app.clear_all()
                        size_hint_x: 0.33
                        height: dp(45)
                        font_size: dp(11)
'''


class TricycleApp(MDApp):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.config_data = []
        self.dialog = None
        self.current_menu = None
        self.selected_model_index = -1

        # 新增后桥品牌、后桥类型默认选项
        self.default_options = {
            'motor_brand': ['巨嵩', '丰利达', '欧比特', '大泰'],
            'motor_power': ['500W', '650W', '800W', '1000W', '1200W', '1500W', '1800W', '2000W', '2200W'],
            'tire_brand': ['东岳', '佳伦', '朝阳', '正新'],
            'front_tire_spec': ['300-12', '350-12', '375-12', '400-12', '450-12'],
            'rear_tire_spec': ['300-12', '350-12', '375-12', '400-12', '450-12', '500-12'],
            'shock_brand': ['睿智', '新超'],
            'shock_model': ['31内簧', '37内簧', '37外簧', '43内簧', '43外簧', '50外簧'],
            'rear_axle_brand': ['常运', '达志美', '伊藤', '大泰'],
            'rear_axle_type': ['分体后桥', '一体后桥', '悬浮后桥']
        }

        # 加载自定义选项
        self.custom_options = {}
        self.load_custom_options()

        # 合并默认选项+自定义选项
        self.options = {}
        for field in self.default_options:
            self.options[field] = self.default_options[field] + self.custom_options.get(field, [])

        # 当前选中配置（新增后桥两项）
        self.current_selections = {
            'motor_brand': '', 'motor_power': '',
            'tire_brand': '', 'front_tire_spec': '', 'rear_tire_spec': '',
            'shock_brand': '', 'shock_model': '',
            'rear_axle_brand': '',
            'rear_axle_type': ''
        }

    def build(self):
        # 中文字体注册
        try:
            LabelBase.register(name="Roboto", fn_regular=r"C:\Windows\Fonts\msyh.ttc")
        except:
            LabelBase.register(name="Roboto", fn_regular=r"C:\Windows\Fonts\simhei.ttf")

        self.theme_cls.theme_style = "Light"
        self.theme_cls.primary_palette = "Blue"
        self.screen = Builder.load_string(KV)
        return self.screen

    def on_start(self):
        self.load_data()
        self.refresh_model_list()

    # 加载自定义选项
    def load_custom_options(self):
        try:
            if os.path.exists('custom_options.json'):
                with open('custom_options.json', 'r', encoding='utf-8') as f:
                    self.custom_options = json.load(f)
        except:
            self.custom_options = {}

    # 保存自定义选项
    def save_custom_options(self):
        try:
            with open('custom_options.json', 'w', encoding='utf-8') as f:
                json.dump(self.custom_options, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f'保存自定义选项失败：{e}')

    # 打开下拉选择菜单
    def open_menu(self, field):
        menu_items = []
        for option in self.options[field]:
            menu_items.append({
                "text": option,
                "viewclass": "OneLineListItem",
                "on_release": lambda x=option: self.select_option(field, x)
            })

        self.current_menu = MDDropdownMenu(
            caller=getattr(self.screen.ids, f'{field}_btn'),
            items=menu_items,
            width_mult=3,
            max_height=dp(200)
        )
        self.current_menu.open()

    # 选择选项
    def select_option(self, field, value):
        self.current_selections[field] = value
        btn = getattr(self.screen.ids, f'{field}_btn')
        btn.text = value
        btn.md_bg_color = (0.2, 0.6, 0.9, 0.3)
        if self.current_menu:
            self.current_menu.dismiss()

    # 手动输入选项弹窗
    def manual_input(self, field, field_name):
        self.dialog = MDDialog(
            title=f'手动输入 - {field_name}',
            type='custom',
            content_cls=MDBoxLayout(
                MDTextField(
                    id='manual_input_field',
                    hint_text=f'请输入{field_name}',
                    mode='rectangle',
                    size_hint_y=None,
                    height=dp(50)
                ),
                orientation='vertical',
                spacing=dp(10),
                size_hint_y=None,
                height=dp(60),
                padding=[dp(10), dp(10)]
            ),
            buttons=[
                MDFlatButton(
                    text='取消',
                    on_release=lambda x: self.dialog.dismiss()
                ),
                MDRaisedButton(
                    text='确定',
                    on_release=lambda x: self.confirm_manual_input(field)
                )
            ]
        )
        self.dialog.open()

    # 确认手动输入
    def confirm_manual_input(self, field):
        content = self.dialog.content_cls
        input_field = content.children[0]
        value = input_field.text.strip()

        if not value:
            self.dialog.dismiss()
            self.show_dialog('提示', '输入内容不能为空！')
            return

        if field not in self.custom_options:
            self.custom_options[field] = []

        if value not in self.custom_options[field] and value not in self.default_options[field]:
            self.custom_options[field].append(value)
            self.save_custom_options()
            self.options[field] = self.default_options[field] + self.custom_options[field]

        self.current_selections[field] = value
        btn = getattr(self.screen.ids, f'{field}_btn')
        btn.text = value
        btn.md_bg_color = (0.8, 0.3, 0.1, 0.3)
        self.dialog.dismiss()

    # 管理自定义选项
    def manage_custom_options(self):
        content = MDBoxLayout(orientation='vertical', spacing=dp(5), size_hint_y=None, height=dp(400))
        scroll = ScrollView()
        list_view = MDList()
        has_custom = False
        field_names = {
            'motor_brand': '电机品牌',
            'motor_power': '电机功率',
            'tire_brand': '轮胎品牌',
            'front_tire_spec': '前轮规格',
            'rear_tire_spec': '后轮规格',
            'shock_brand': '减震品牌',
            'shock_model': '减震型号',
            'rear_axle_brand': '后桥品牌',
            'rear_axle_type': '后桥类型'
        }
        for field, options in self.custom_options.items():
            if options:
                has_custom = True
                list_view.add_widget(OneLineListItem(text=f'【{field_names.get(field, field)}】'))
                for option in options:
                    item = TwoLineListItem(text=f'  • {option}', secondary_text='点击删除此选项',
                                           on_release=lambda x, f=field, o=option: self.delete_custom_option(f, o))
                    list_view.add_widget(item)
        if not has_custom:
            list_view.add_widget(OneLineListItem(text='暂无自定义选项'))
        scroll.add_widget(list_view)
        content.add_widget(scroll)
        self.dialog = MDDialog(title='管理自定义选项', type='custom', content_cls=content,
                               buttons=[MDFlatButton(text='关闭', on_release=lambda x: self.dialog.dismiss())],
                               size_hint=(0.9, 0.7))
        self.dialog.open()

    # 删除自定义选项
    def delete_custom_option(self, field, option):
        if field in self.custom_options and option in self.custom_options[field]:
            self.custom_options[field].remove(option)
            self.save_custom_options()
            self.options[field] = self.default_options[field] + self.custom_options[field]
            self.dialog.dismiss()
            self.show_dialog('已删除', f'已删除选项：{option}')
            self.manage_custom_options()

    # 重置所有选择
    def reset_selections(self):
        for field in self.current_selections:
            self.current_selections[field] = ''
            btn = getattr(self.screen.ids, f'{field}_btn')
            btn.text = '点击选择 ▼'
            btn.md_bg_color = (0.9, 0.9, 0.9, 1)

    # 添加配置到列表（修复重复校验，允许同名不同配置）
    def add_to_list(self):
        model_name = self.screen.ids.model_input.text.strip()
        price_text = self.screen.ids.price_input.text.strip()
        quantity_text = self.screen.ids.quantity_input.text.strip()

        if not model_name:
            self.show_dialog('提示', '请输入车型号！')
            return
        if not price_text or not quantity_text:
            self.show_dialog('提示', '请输入单价和数量！')
            return

        try:
            price = float(price_text)
            quantity = int(quantity_text)
            if price <= 0 or quantity <= 0:
                raise ValueError
        except ValueError:
            self.show_dialog('提示', '请输入有效的数字！')
            return

        # 校验字段：包含新增后桥两项
        check_fields = [
            '电机品牌','电机功率','轮胎品牌','前轮规格','后轮规格',
            '减震品牌','减震型号','rear_axle_brand','rear_axle_type'
        ]
        # 遍历已有数据做重复校验
        for item in self.config_data:
            if item['车型号'] == model_name:
                same_config = True
                for field in check_fields:
                    old_val = item.get(field, '-')
                    new_val = self.current_selections.get(field, '-')
                    if old_val != new_val:
                        same_config = False
                        break
                if same_config:
                    self.show_dialog('提示', f'该车型"{model_name}"的这套配置已经存在，请勿重复添加！')
                    return

        # 保存数据（新增后桥字段）
        config = {
            '车型号': model_name,
            '电机品牌': self.current_selections['motor_brand'] or '-',
            '电机功率': self.current_selections['motor_power'] or '-',
            '轮胎品牌': self.current_selections['tire_brand'] or '-',
            '前轮规格': self.current_selections['front_tire_spec'] or '-',
            '后轮规格': self.current_selections['rear_tire_spec'] or '-',
            '减震品牌': self.current_selections['shock_brand'] or '-',
            '减震型号': self.current_selections['shock_model'] or '-',
            '后桥品牌': self.current_selections['rear_axle_brand'] or '-',
            '后桥类型': self.current_selections['rear_axle_type'] or '-',
            '单价(元)': price,
            '数量(辆)': quantity,
            '小计(元)': price * quantity,
            '备注': self.screen.ids.remarks_input.text.strip() or '-',
            '录入时间': datetime.now().strftime('%m-%d %H:%M')
        }

        self.config_data.append(config)
        self.refresh_model_list()
        self.update_stats()
        self.save_data()

        selected_count = sum(1 for v in self.current_selections.values() if v)
        self.show_dialog('成功', f'已添加：{model_name}\n已选{selected_count}项配置\n数量：{quantity}辆 | ¥{price * quantity:,.2f}')

    # 快速录入下一个
    def quick_add(self):
        model_name = self.screen.ids.model_input.text.strip()
        if not model_name:
            self.show_dialog('提示', '请输入车型号！')
            return
        self.add_to_list()
        if any(item['车型号'] == model_name for item in self.config_data):
            self.screen.ids.model_input.text = ''
            self.screen.ids.price_input.text = ''
            self.screen.ids.quantity_input.text = ''
            self.screen.ids.remarks_input.text = ''
            self.reset_selections()
            self.screen.ids.model_input.focus = True

    # 刷新型号列表
    def refresh_model_list(self):
        list_widget = self.screen.ids.model_list_preview
        list_widget.clear_widgets()
        for i, config in enumerate(self.config_data):
            sub_text = f"数量:{config['数量(辆)']}辆 | ¥{config['单价(元)']:,.0f}"
            if config['电机品牌'] != '-':
                sub_text += f" | {config['电机品牌']}"
            item = TwoLineListItem(text=f"{i + 1}. {config['车型号']}", secondary_text=sub_text,
                                   on_release=lambda x, idx=i: self.select_model(idx))
            if i == self.selected_model_index:
                item.bg_color = (0.2, 0.6, 0.9, 0.2)
            list_widget.add_widget(item)
        self.screen.ids.count_label.text = f'{len(self.config_data)}个'

    # 查看型号详情
    def select_model(self, index):
        if 0 <= index < len(self.config_data):
            self.selected_model_index = index
            self.refresh_model_list()
            config = self.config_data[index]
            info = (
                f"车型号：{config['车型号']}\n"
                f"电机：{config['电机品牌']} {config['电机功率']}\n"
                f"轮胎：{config['轮胎品牌']}\n"
                f"前轮：{config['前轮规格']} | 后轮：{config['后轮规格']}\n"
                f"减震：{config['减震品牌']} {config['减震型号']}\n"
                f"后桥：{config['后桥品牌']} {config['后桥类型']}\n"
                f"数量：{config['数量(辆)']}辆 | ¥{config['小计(元)']:,.2f}"
            )
            self.show_dialog('型号详情', info)

    # 更新统计栏
    def update_stats(self):
        total_models = len(self.config_data)
        total_quantity = sum(item['数量(辆)'] for item in self.config_data)
        total_amount = sum(item['小计(元)'] for item in self.config_data)
        self.screen.ids.stats_bar.text = f'已录入: {total_models}个型号 | 总计: {total_quantity}辆 | 总金额: ¥{total_amount:,.2f}'

    # 查看完整表格
    def view_full_table(self):
        if not self.config_data:
            self.show_dialog('提示', '暂无数据')
            return
        table_data = []
        for i, config in enumerate(self.config_data, 1):
            table_data.append([
                str(i), config['车型号'], config['电机品牌'], config['电机功率'],
                config['轮胎品牌'], config['前轮规格'], config['后轮规格'],
                config['减震品牌'], config['减震型号'], config['后桥品牌'], config['后桥类型'],
                f"¥{config['单价(元)']:,.0f}", str(config['数量(辆)']), f"¥{config['小计(元)']:,.0f}", config['备注']
            ])
        total_quantity = sum(item['数量(辆)'] for item in self.config_data)
        total_amount = sum(item['小计(元)'] for item in self.config_data)
        table_data.append(['', '【汇总】', '', '', '', '', '', '', '', '', '', str(total_quantity), f"¥{total_amount:,.0f}", ''])
        column_data = [
            ("序号", dp(10)), ("车型号", dp(22)), ("电机品牌", dp(15)), ("电机功率", dp(15)),
            ("轮胎品牌", dp(15)), ("前轮规格", dp(15)), ("后轮规格", dp(15)),
            ("减震品牌", dp(15)), ("减震型号", dp(15)), ("后桥品牌", dp(15)), ("后桥类型", dp(15)),
            ("单价", dp(15)), ("数量", dp(12)), ("小计", dp(15)), ("备注", dp(25))
        ]
        content = MDBoxLayout(orientation='vertical', spacing=dp(5), size_hint_y=None, height=dp(520))
        data_table = MDDataTable(size_hint=(1, 0.88), use_pagination=True, rows_num=10, column_data=column_data,
                                 row_data=table_data, elevation=2)
        content.add_widget(data_table)
        btn_box = MDBoxLayout(orientation='horizontal', spacing=dp(10), size_hint_y=None, height=dp(45),
                              padding=[dp(10), dp(5)])
        btn_box.add_widget(MDRaisedButton(text='关闭', size_hint_x=0.5, on_release=lambda x: self.dialog.dismiss()))
        btn_box.add_widget(MDRaisedButton(text='导出Excel', md_bg_color=(0.2, 0.7, 0.3, 1), size_hint_x=0.5,
                                          on_release=lambda x: self.export_and_close()))
        content.add_widget(btn_box)
        self.dialog = MDDialog(title=f'配置总表 (共{len(self.config_data)}个型号)', type='custom', content_cls=content,
                               size_hint=(0.96, 0.88))
        self.dialog.open()

    def export_and_close(self):
        self.dialog.dismiss()
        self.export_excel()

    # 修复Excel导出（解决合并单元格报错、保存到程序文件夹）
    def export_excel(self):
        if not self.config_data:
            self.show_dialog('提示', '没有数据可导出')
            return
        try:
            export_data = []
            for i, config in enumerate(self.config_data, 1):
                export_data.append({
                    '序号': i, '车型号': config['车型号'], '电机品牌': config['电机品牌'],
                    '电机功率': config['电机功率'], '轮胎品牌': config['轮胎品牌'],
                    '前轮规格': config['前轮规格'], '后轮规格': config['后轮规格'],
                    '减震品牌': config['减震品牌'], '减震型号': config['减震型号'],
                    '后桥品牌': config['后桥品牌'], '后桥类型': config['后桥类型'],
                    '单价(元)': config['单价(元)'], '数量(辆)': config['数量(辆)'],
                    '小计(元)': config['小计(元)'], '备注': config['备注'], '录入时间': config['录入时间']
                })
            df = pd.DataFrame(export_data)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f'三轮车配置表_{timestamp}.xlsx'
            # 保存到当前脚本文件夹
            save_path = os.path.join(os.getcwd(), filename)
            print("Excel保存路径：", save_path)

            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='配置清单', index=False, startrow=2)
                worksheet = writer.sheets['配置清单']
                worksheet.cell(row=1, column=1, value=f'三轮车配置清单 - {datetime.now().strftime("%Y年%m月%d日")}')
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

                # 修复合并单元格列宽报错
                from openpyxl.cell.cell import Cell
                for col in worksheet.columns:
                    first_cell = col[0]
                    if not isinstance(first_cell, Cell):
                        continue
                    col_letter = first_cell.column_letter
                    cell_list = [cell for cell in col if isinstance(cell, Cell)]
                    max_len = max(len(str(cell.value or '')) for cell in cell_list)
                    worksheet.column_dimensions[col_letter].width = min(max_len + 2, 40)

                last_row = len(df) + 4
                from openpyxl.styles import Font, PatternFill
                worksheet.cell(row=last_row, column=1, value='汇总')
                worksheet.cell(row=last_row, column=13, value=df['数量(辆)'].sum())
                worksheet.cell(row=last_row, column=14, value=df['小计(元)'].sum())
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=last_row, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

            total_qty = df['数量(辆)'].sum()
            total_amt = df['小计(元)'].sum()
            self.show_dialog('导出成功', f'文件已保存在程序文件夹内\n路径：{save_path}\n{len(self.config_data)}个型号\n{total_qty}辆车\n总金额：¥{total_amt:,.2f}')
        except Exception as e:
            import traceback
            err = traceback.format_exc()
            print("导出异常：", err)
            self.show_dialog('导出失败', f'错误信息：{str(e)}')

    # 删除选中
    def delete_selected(self):
        if self.selected_model_index < 0 or self.selected_model_index >= len(self.config_data):
            self.show_dialog('提示', '请先点击列表中的型号')
            return
        config = self.config_data[self.selected_model_index]
        self.dialog = MDDialog(
            title='确认删除',
            text=f'删除型号：{config["车型号"]}？\n数量：{config["数量(辆)"]}辆\n小计：¥{config["小计(元)"]:,.2f}',
            buttons=[
                MDFlatButton(text='取消', on_release=lambda x: self.dialog.dismiss()),
                MDRaisedButton(text='删除', md_bg_color=(0.9, 0.2, 0.2, 1),
                               on_release=lambda x: self.do_delete_selected())
            ]
        )
        self.dialog.open()

    def do_delete_selected(self):
        if 0 <= self.selected_model_index < len(self.config_data):
            deleted = self.config_data.pop(self.selected_model_index)
            self.selected_model_index = -1
            self.refresh_model_list()
            self.update_stats()
            self.save_data()
            self.dialog.dismiss()
            self.show_dialog('已删除', f'型号：{deleted["车型号"]}')

    # 删除最后一条
    def delete_last(self):
        if not self.config_data:
            self.show_dialog('提示', '没有数据')
            return
        last = self.config_data[-1]
        self.dialog = MDDialog(
            title='确认删除', text=f'删除：{last["车型号"]}？',
            buttons=[
                MDFlatButton(text='取消', on_release=lambda x: self.dialog.dismiss()),
                MDRaisedButton(text='删除', md_bg_color=(0.9, 0.5, 0.1, 1), on_release=lambda x: self.do_delete_last())
            ]
        )
        self.dialog.open()

    def do_delete_last(self):
        if self.config_data:
            deleted = self.config_data.pop()
            if self.selected_model_index >= len(self.config_data):
                self.selected_model_index = -1
            self.refresh_model_list()
            self.update_stats()
            self.save_data()
            self.dialog.dismiss()
            self.show_dialog('已删除', f'型号：{deleted["车型号"]}')

    # 清空全部
    def clear_all(self):
        if not self.config_data:
            self.show_dialog('提示', '没有数据')
            return
        self.dialog = MDDialog(
            title='确认清空', text=f'删除全部{len(self.config_data)}个型号？',
            buttons=[
                MDFlatButton(text='取消', on_release=lambda x: self.dialog.dismiss()),
                MDRaisedButton(text='确认', md_bg_color=(0.8, 0.1, 0.1, 1), on_release=lambda x: self.do_clear_all())
            ]
        )
        self.dialog.open()

    def do_clear_all(self):
        self.config_data = []
        self.selected_model_index = -1
        self.refresh_model_list()
        self.update_stats()
        self.save_data()
        self.dialog.dismiss()
        self.show_dialog('完成', '所有数据已清空')

    # 保存数据
    def save_data(self):
        try:
            with open('tricycle_choice_data.json', 'w', encoding='utf-8') as f:
                json.dump(self.config_data, f, ensure_ascii=False, indent=2)
        except:
            pass

    # 加载数据（自动补齐所有新字段，防止KeyError闪退）
    def load_data(self):
        try:
            if os.path.exists('tricycle_choice_data.json'):
                with open('tricycle_choice_data.json', 'r', encoding='utf-8') as f:
                    self.config_data = json.load(f)
                # 补齐所有新增字段
                fill_fields = [
                    '电机品牌','电机功率','轮胎品牌','前轮规格','后轮规格',
                    '减震品牌','减震型号','后桥品牌','后桥类型'
                ]
                for row in self.config_data:
                    for field in fill_fields:
                        if field not in row:
                            row[field] = '-'
                self.update_stats()
        except:
            self.config_data = []

    # 通用弹窗
    def show_dialog(self, title, text):
        if self.dialog:
            self.dialog.dismiss()
        self.dialog = MDDialog(
            title=title, text=text, size_hint=(0.85, 0.25),
            buttons=[MDFlatButton(text='确定', on_release=lambda x: self.dialog.dismiss())]
        )
        self.dialog.open()


if __name__ == '__main__':
    TricycleApp().run()